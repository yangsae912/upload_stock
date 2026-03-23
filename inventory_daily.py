import os
import psycopg2
import openpyxl
from datetime import date
import sys

# DB 연결 설정 (환경에 맞게 수정)
DB_CONFIG = {
    "host": os.getenv("PGHOST", ""),
    "database": os.getenv("PGDATABASE", "pos"),
    "user": os.getenv("PGUSER", "postgres"),
    "password": os.getenv("PGPASSWORD", "postgres"),
    "port": int(os.getenv("PGPORT", "5432"))
}

def load_inventory(file_path: str, snapshot_date: date = None):
    if snapshot_date is None:
        snapshot_date = date.today()

    # xlsx 파싱
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    def to_int(val):
        try:
            return int(str(val).replace(',', '').strip() or 0)
        except:
            return 0

    def to_bool(val):
        return str(val).strip() in ('1', 'true', 'True', 'Y', 'y', '품절')

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        hcode = str(row[idx['옵션추가항목1']] or '').strip()

        # 빈 값 제외
        if not hcode:
            continue

        normal_stock    = to_int(row[idx['정상재고']])
        defect_stock    = to_int(row[idx['불량재고']])
        available_stock = normal_stock - defect_stock
        incoming_stock  = to_int(row[idx['입고대기']])
        is_soldout      = to_bool(row[idx['품절']])

        records.append((
            snapshot_date,
            hcode,
            str(row[idx['상품코드']] or '').strip(),
            str(row[idx['상품명']] or '').strip(),
            str(row[idx['옵션']] or '').strip(),
            str(row[idx['카테고리']] or '').strip(),
            normal_stock,
            defect_stock,
            available_stock,
            incoming_stock,
            is_soldout,
            str(row[idx['옵션추가항목2']] or '').strip(),   # 추가

        ))

    wb.close()
    print(f"파싱 완료: {len(records)}개 상품")

    # DB 적재
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # 당일 데이터 삭제 후 재적재 (멱등성 보장)
    cur.execute("DELETE FROM public.inventory_daily WHERE snapshot_date = %s", (snapshot_date,))

    cur.executemany("""
        INSERT INTO public.inventory_daily (
            snapshot_date, hcode, sku_code, product_name, option_name, category,
            normal_stock, defect_stock, available_stock, incoming_stock, is_soldout, product_grade
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, records)

    conn.commit()
    cur.close()
    conn.close()

    print(f"적재 완료: {len(records)}개 상품 ({snapshot_date})")

if __name__ == "__main__":
    file_path = sys.argv[1] if len(sys.argv) > 1 else "재고조회.xlsx"
    snapshot_date = date.fromisoformat(sys.argv[2]) if len(sys.argv) > 2 else date.today()
    load_inventory(file_path, snapshot_date)